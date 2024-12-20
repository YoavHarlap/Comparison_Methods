import fitz
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def extract_pdf_annotations_to_excel_with_formatting(pdf_path, excel_path):
    annotations = []
    try:
        # Open the PDF document
        pdf_document = fitz.open(pdf_path)

        # Iterate through each page
        for page_num in range(len(pdf_document)):
            page = pdf_document[page_num]

            # Retrieve annotations on the page
            for annot in page.annots():
                # Extract details of the annotation
                annot_rect = annot.rect  # Get the rectangle of the annotation
                nearby_text = page.get_text("words")  # Extract words from the page

                # Find words near the annotation rectangle
                matched_words = [
                    word[4] for word in nearby_text
                    if fitz.Rect(word[:4]).intersects(annot_rect)
                ]

                # Combine matched words into a single string
                matched_text = " ".join(matched_words) if matched_words else "No matching words"

                annot_info = {
                    "Page": page_num + 1,  # Page numbers are 1-based
                    "Annotation Text": annot.info.get("content", "No content"),
                    "Matched Words": matched_text
                }
                annotations.append(annot_info)

        # Close the document
        pdf_document.close()

        # Convert the annotations list to a pandas DataFrame
        annotations_df = pd.DataFrame(annotations)

        # Save the DataFrame to an Excel file
        annotations_df.to_excel(excel_path, index=False)

        # Open the Excel file with openpyxl for formatting
        workbook = load_workbook(excel_path)
        sheet = workbook.active

        # Auto-adjust column widths
        for col in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = max_length + 2  # Add some padding
            sheet.column_dimensions[col_letter].width = adjusted_width

        # Wrap text for all cells
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = cell.alignment.copy(wrap_text=True)

        # Adjust row heights based on content
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value:
                    lines = str(cell.value).count("\n") + 1
                    sheet.row_dimensions[cell.row].height = lines * 15  # Adjust row height

        # Save the formatted Excel file
        workbook.save(excel_path)
        print(f"Annotations saved and formatted to Excel at: {excel_path}")
    except Exception as e:
        print(f"Error reading PDF or saving to Excel: {e}")


# Path to your PDF file
pdf_file = r"C:\Users\ASUS\Documents\code_images\תזה גרסאות\3\Phase_retrieval_and_matrix_completion_through_projection_based_algorithms tamir.pdf"
# Path to save the Excel file
excel_file = r"C:\Users\ASUS\Documents\code_images\תזה גרסאות\3\PDF_Annotations_Formatted.xlsx"

# Extract annotations, save, and format Excel
extract_pdf_annotations_to_excel_with_formatting(pdf_file, excel_file)
